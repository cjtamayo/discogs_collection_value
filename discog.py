from datetime import datetime
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font
import re
import requests
import time


def timer(func):
    """
    Print the runtime of the decorated function
    :param func: function that we want to be timed
    :return: value of function, but prints string of how long function ran
    """
    import functools
    import time
    @functools.wraps(func)
    def wrapper_timer(*args, **kwargs):
        start_time = time.time()
        value = func(*args, **kwargs)
        end_time = time.time()
        run_time = end_time - start_time
        if run_time > 3659:
            hours = int(run_time/3600)
            print(f"Finished {func.__name__} in {hours:.0f} hours, {(run_time-hours*3600)/60:.0f} minutes and {run_time%60:.0f} secs")
        elif run_time > 59:
            print(f"Finished {func.__name__} in {run_time/60:.0f} minutes and {run_time%60:.0f} secs")
        else:
            print(f"Finished {func.__name__} in {run_time:.2f} secs")
        return value

    return wrapper_timer


def collection_call(next_page=1, page_call=''):
    """
    Makes an API call to get collection
    :param next_page: boolean if this is the first call
    :param string of api call
    :return: list of details for collection item, pagination dict
    """
    user_name = os.getenv('USER_NAME')
    token = os.getenv('DISCOGS_TOKEN')
    if next_page == 1:
        get_url = 'https://api.discogs.com/users/{}/collection/folders/0/releases?token={}'.format(user_name, token)
    else:
        get_url = page_call

    r = requests.get(get_url)
    response = r.json()

    return response.get("releases"), response.get("pagination")


def collection_grab():
    """
    Iterates over discogs collection API call to build a list() of titles
    :return: completed list of all items in Discogs collection
    """
    titles = list()

    init_titles, pagination = collection_call()
    titles += init_titles

    page_tries = pagination.get("pages")
    next_call = pagination.get("urls", {}).get("next")

    if page_tries == 1:
        print("Titles less than 50")
    else:
        for page in range(2, page_tries+1):
            new_titles, new_page = collection_call(page, next_call)
            titles += new_titles
            next_call = new_page.get("urls", {}).get("next")

    print("Total collection is {} items".format(len(titles)))

    return titles


def title_flatten(title_dict):
    """
    Flattens the title dictionary, removing some details
    :param title_dict: dict() of json title response from Discogs
    :return:
    """
    time.sleep(2)
    title_flat = dict()
    bi = title_dict.get("basic_information")
    title_id = title_dict.get("id")
    discog_artist = bi.get("artists", {})[0]["name"]
    artist = re.sub('.(\d+\))', '', discog_artist).rstrip()
    title = bi.get("title")

    # IDs

    title_flat["id"] = title_id
    title_flat["master_id"] = bi.get("master_id")

    # Time
    date_long = title_dict.get("date_added").split("T")
    title_flat["date_added"] = date_long[0]
    title_flat["time_added"] = date_long[1]

    # Release Details
    title_flat["artist"] = artist
    title_flat["artist_id"] = bi.get("artists", {})[0]["id"]
    title_flat["title"] = title
    title_flat["year"] = bi.get("year")
    title_flat["format"] = bi.get("formats", {})[0]["name"]
    title_flat["format_info"] = bi.get("formats", {})[0].get("text", "N/A")
    title_flat["genres"] = str(bi.get("genres"))[1:-1]
    title_flat["styles"] = str(bi.get("styles"))[1:-1]

    # Collection Notes
    title_flat["notes"] = title_dict.get("notes", {})[0]["value"]

    # Get Lowest Price
    rs_url = requests.get('https://api.discogs.com/marketplace/stats/{}?curr_abbr=USD&token={}'.format(title_id,
                                                                                    os.getenv('DISCOGS_TOKEN')))
    release_stats = rs_url.json()
    if not release_stats.get("num_for_sale"):
        print("{} by {} has no copies for sale".format(title, artist))
        lowest_price = 0.0
    else:
        lowest_price = release_stats.get("lowest_price", {}).get("value", 0.0)

    title_flat["lowest_price"] = lowest_price

    return title_flat


def collection_lowest_price(title_list):
    """
    Prints out total value (based on lowest available price in market) for your collection
    :param title_list:
    :return:
    """
    missing = list()
    prices = list()

    for title in title_list:
        price = title.get("lowest_price")
        if not price:
            missing.append(title.get("title"))
        else:
            prices.append(price)

    tot_value = round(sum(prices), 2)
    str_missing = str(missing)[1:-1]

    print("Total value is {} but prices are missing from {}".format(tot_value, str_missing))

    return


def list_to_xl(title_list):
    """
    Iterates flattend lists and converts them to excel, saved in current folder
    :param title_list:
    :return:
    """
    # Creating Workbook
    wb = Workbook()
    ws = wb.active

    # Setting Up Headers
    ws['A1'] = 'ID'
    ws['B1'] = 'Artist'
    ws['C1'] = 'Title'
    ws['D1'] = 'Year'
    ws['E1'] = 'Format'
    ws['F1'] = 'Format Info'
    ws['G1'] = 'Genres'
    ws['H1'] = 'Styles'
    ws['I1'] = 'Lowest Price'
    ws['J1'] = 'Date Added'
    ws['K1'] = 'Time Added'
    ws['L1'] = 'Artist ID'
    ws['M1'] = 'Master ID'
    ws['N1'] = 'Notes'

    a1 = ws['A1']
    b1 = ws['B1']
    c1 = ws['C1']
    d1 = ws['D1']
    e1 = ws['E1']
    f1 = ws['F1']
    g1 = ws['G1']
    h1 = ws['H1']
    i1 = ws['I1']
    j1 = ws['J1']
    k1 = ws['K1']
    l1 = ws['L1']
    m1 = ws['M1']
    n1 = ws['N1']

    headers = [a1, b1, c1, d1, e1, f1, g1, h1, i1, j1, k1, l1, m1, n1]
    for cell in headers:
        cell.font = Font(size=14, bold=True)
    for row_num, row in enumerate(title_list, 2):
        ws['A{}'.format(row_num)] = row.get("id")
        ws['B{}'.format(row_num)] = row.get('artist')
        ws['C{}'.format(row_num)] = row.get('title')
        ws['D{}'.format(row_num)] = row.get('year')
        ws['E{}'.format(row_num)] = row.get('format')
        ws['F{}'.format(row_num)] = row.get('format_info')
        ws['G{}'.format(row_num)] = row.get('genres')
        ws['H{}'.format(row_num)] = row.get('styles')
        ws['I{}'.format(row_num)] = row.get('lowest_price')
        ws['J{}'.format(row_num)] = row.get('date_added')
        ws['K{}'.format(row_num)] = row.get('time_added')
        ws['L{}'.format(row_num)] = row.get('artist_id')
        ws['M{}'.format(row_num)] = row.get('master_id')
        ws['N{}'.format(row_num)] = row.get('notes')

    # Saving
    today = str(datetime.today())[:10].replace('-','_')
    wb.save("discogs_collection{}.xlsx".format(today))

    return








test_listo = [{'id': 7816932, 'master_id': 910103, 'date_added': '2017-05-05', 'time_added': '21:34:15-07:00', 'artist': 'Grimes', 'artist_id': 1993487, 'title': 'Art Angels', 'year': 2015, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Electronic', 'Pop'", 'styles': "'Synth-pop', 'Experimental'", 'notes': 'Purchased on Amazon, 2016', 'lowest_price': 14.99}, {'id': 3419668, 'master_id': 411575, 'date_added': '2017-05-05', 'time_added': '21:50:56-07:00', 'artist': 'Grimes', 'artist_id': 1993487, 'title': 'Visions', 'year': 2012, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Electronic', 'Pop'", 'styles': "'Leftfield', 'New Wave', 'Abstract', 'Synth-pop', 'Experimental', 'Ethereal'", 'notes': 'Purchased at Turntable Lab, Manhattan, RSD 2017', 'lowest_price': 12.99}, {'id': 5472807, 'master_id': 7882, 'date_added': '2017-06-06', 'time_added': '20:30:47-07:00', 'artist': 'Slayer', 'artist_id': 18845, 'title': 'God Hates Us All', 'year': 2014, 'format': 'Vinyl', 'format_info': 'Clear, 180g', 'genres': "'Rock'", 'styles': "'Thrash'", 'notes': '#503/666\nPurchased on Nuclear Blast, 2017', 'lowest_price': 111.76}, {'id': 8062757, 'master_id': 370255, 'date_added': '2017-05-05', 'time_added': '21:32:30-07:00', 'artist': 'Grimes', 'artist_id': 1993487, 'title': 'Geidi Primes', 'year': 2016, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Electronic', 'Pop'", 'styles': "'Synth-pop', 'Experimental'", 'notes': 'Purchased at Turntable Lab, Manhattan, RSD 2017', 'lowest_price': 10.39}, {'id': 1079745, 'master_id': 40617, 'date_added': '2017-05-05', 'time_added': '21:16:12-07:00', 'artist': 'James Blood Ulmer', 'artist_id': 149965, 'title': 'Are You Glad To Be In America?', 'year': 1981, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Jazz', 'Funk / Soul'", 'styles': "'Free Jazz', 'Contemporary Jazz', 'Funk'", 'notes': 'Free gift at Bleeker Street Records, Manhattan, RSD 2016', 'lowest_price': 11.27}, {'id': 8396148, 'master_id': 987819, 'date_added': '2017-05-05', 'time_added': '21:22:50-07:00', 'artist': 'Iron Maiden', 'artist_id': 251595, 'title': 'Empire Of The Clouds', 'year': 2016, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock', 'Pop'", 'styles': "'Ballad'", 'notes': 'Purchased at Bleeker Street Records, Manhattan, RSD 2016', 'lowest_price': 34.99}, {'id': 9144760, 'master_id': 71534, 'date_added': '2017-05-22', 'time_added': '17:58:39-07:00', 'artist': 'The Killers', 'artist_id': 220651, 'title': "Sam's Town", 'year': 2016, 'format': 'Vinyl', 'format_info': '10th Anniversary Edition', 'genres': "'Electronic', 'Rock'", 'styles': "'New Wave', 'Indie Rock', 'Glam'", 'notes': '# 04947/05016\nPurchased on Bong Load, 2017', 'lowest_price': 47.06}, {'id': 749903, 'master_id': 37674, 'date_added': '2017-05-05', 'time_added': '19:19:58-07:00', 'artist': 'Bob Seger And The Silver Bullet Band', 'artist_id': 268911, 'title': 'Stranger In Town', 'year': 1978, 'format': 'Vinyl', 'format_info': 'Winchester Press', 'genres': "'Rock'", 'styles': "'Rock & Roll', 'Soft Rock', 'Classic Rock'", 'notes': 'Purchased at A1 Record Shop, Manhattan, 2017', 'lowest_price': 2.06}, {'id': 6255193, 'master_id': 41722, 'date_added': '2017-06-08', 'time_added': '19:13:49-07:00', 'artist': 'Kiss', 'artist_id': 153073, 'title': 'Destroyer', 'year': 1977, 'format': 'Vinyl', 'format_info': 'Santa Maria Pressing', 'genres': "'Rock'", 'styles': "'Hard Rock'", 'notes': 'Purchased at A1 Record Shop, Manhattan, 2017', 'lowest_price': 5.99}, {'id': 8513748, 'master_id': 41191, 'date_added': '2017-05-05', 'time_added': '12:05:14-07:00', 'artist': 'Ozzy Osbourne', 'artist_id': 59770, 'title': 'The Ultimate Sin', 'year': 1986, 'format': 'Vinyl', 'format_info': 'Pitman Pressing', 'genres': "'Rock'", 'styles': "'Heavy Metal'", 'notes': 'Purchased at A1 Record Shop, Manhattan, 2017', 'lowest_price': 29.41}, {'id': 7747307, 'master_id': 41148, 'date_added': '2017-05-05', 'time_added': '20:51:04-07:00', 'artist': 'Ozzy Osbourne', 'artist_id': 59770, 'title': 'Bark At The Moon', 'year': 1983, 'format': 'Vinyl', 'format_info': 'Pitman Press', 'genres': "'Rock'", 'styles': "'Heavy Metal'", 'notes': 'Purchased at A1 Record Shop, Manhattan, 2017', 'lowest_price': 14.99}, {'id': 412816, 'master_id': 26341, 'date_added': '2017-06-08', 'time_added': '20:43:32-07:00', 'artist': 'Judas Priest', 'artist_id': 252121, 'title': 'Screaming For Vengeance', 'year': 1982, 'format': 'Vinyl', 'format_info': 'Carrollton Pressing', 'genres': "'Rock'", 'styles': "'Heavy Metal', 'Hard Rock'", 'notes': 'Purchased at A1 Record Shop, Manhattan, 2017', 'lowest_price': 8.99}, {'id': 707930, 'master_id': 27757, 'date_added': '2017-05-05', 'time_added': '08:42:38-07:00', 'artist': 'Simon & Garfunkel', 'artist_id': 232157, 'title': 'Bookends', 'year': 1968, 'format': 'Vinyl', 'format_info': 'Santa Maria Press', 'genres': "'Rock', 'Pop'", 'styles': "'Folk Rock', 'Soft Rock', 'Pop Rock', 'Vocal'", 'notes': 'Purchased at Good Records NYC, Manhattan, 2017', 'lowest_price': 0.99}, {'id': 1446879, 'master_id': 37598, 'date_added': '2017-06-08', 'time_added': '19:32:52-07:00', 'artist': 'Santana', 'artist_id': 30724, 'title': 'Moonflower', 'year': 1977, 'format': 'Vinyl', 'format_info': 'Pitman Pressing', 'genres': "'Jazz', 'Rock'", 'styles': "'Fusion', 'Classic Rock', 'Latin Jazz'", 'notes': 'Purchased at A1 Record Shop, Manhattan, 2017', 'lowest_price': 2.24}, {'id': 10244877, 'master_id': 68791, 'date_added': '2017-05-05', 'time_added': '19:53:53-07:00', 'artist': 'Billy Joel', 'artist_id': 137418, 'title': 'Glass Houses', 'year': 1980, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock', 'Pop'", 'styles': "'Pop Rock', 'Rock & Roll'", 'notes': 'Purchased at A1 Record Shop, Manhattan, 2017', 'lowest_price': 2.95}, {'id': 1256010, 'master_id': 72126, 'date_added': '2017-06-08', 'time_added': '20:18:41-07:00', 'artist': 'Billy Joel', 'artist_id': 137418, 'title': 'The Bridge', 'year': 1986, 'format': 'Vinyl', 'format_info': 'Pitman', 'genres': "'Rock', 'Blues', 'Pop'", 'styles': "'Piano Blues', 'Pop Rock', 'Ballad'", 'notes': 'Purchased at A1 Record Shop, Manhattan, 2017', 'lowest_price': 0.98}, {'id': 10243021, 'master_id': 57048, 'date_added': '2017-05-05', 'time_added': '11:12:57-07:00', 'artist': 'Billy Joel', 'artist_id': 137418, 'title': '52nd Street', 'year': 0, 'format': 'Vinyl', 'format_info': 'Carrollton', 'genres': "'Jazz', 'Rock', 'Pop'", 'styles': "'Pop Rock', 'Ballad', 'Latin Jazz'", 'notes': 'Purchased at Academy Records, Manhattan, 2017', 'lowest_price': 1.18}, {'id': 10245007, 'master_id': 57029, 'date_added': '2017-05-05', 'time_added': '20:23:00-07:00', 'artist': 'Billy Joel', 'artist_id': 137418, 'title': 'The Stranger', 'year': 1977, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': "'Pop Rock'", 'notes': 'Purchased at A1 Record Shop, Manhattan, 2017', 'lowest_price': 6.0}, {'id': 10168661, 'master_id': 1170192, 'date_added': '2017-05-05', 'time_added': '22:00:30-07:00', 'artist': 'Santana', 'artist_id': 30724, 'title': 'Woodstock (Saturday, August 16, 1969)', 'year': 2017, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock', 'Latin'", 'styles': '', 'notes': 'Purchased at Turntable Lab, Manhattan, RSD 2017', 'lowest_price': 60.0}, {'id': 10175932, 'master_id': 468538, 'date_added': '2017-05-05', 'time_added': '21:52:36-07:00', 'artist': 'Robert Johnson', 'artist_id': 272142, 'title': 'The Centennial Collection', 'year': 2017, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Blues'", 'styles': "'Delta Blues'", 'notes': 'Purchased at Turntable Lab, Manhattan, RSD 2017', 'lowest_price': 110.0}, {'id': 2055654, 'master_id': 21666, 'date_added': '2017-07-10', 'time_added': '16:59:10-07:00', 'artist': 'Alice In Chains', 'artist_id': 251846, 'title': 'Dirt', 'year': 2009, 'format': 'Vinyl', 'format_info': '180 Gram', 'genres': "'Rock'", 'styles': "'Alternative Rock', 'Hard Rock', 'Grunge'", 'notes': 'Purchased on Discogs, 2017', 'lowest_price': 20.59}, {'id': 3249944, 'master_id': 0, 'date_added': '2017-05-05', 'time_added': '21:59:00-07:00', 'artist': 'Pink Floyd', 'artist_id': 45467, 'title': 'The Wall Singles Collection', 'year': 2011, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': "'Prog Rock'", 'notes': 'Purchased at J&R, Manhattan, RSD 2012', 'lowest_price': 64.71}, {'id': 3580556, 'master_id': 0, 'date_added': '2017-05-05', 'time_added': '22:05:43-07:00', 'artist': 'Ozzy Osbourne', 'artist_id': 59770, 'title': 'Believer (Live)', 'year': 2012, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': "'Hard Rock', 'Heavy Metal'", 'notes': 'Purchased at J&R, Manhattan, RSD 2012', 'lowest_price': 22.62}, {'id': 2414643, 'master_id': 67090, 'date_added': '2017-05-05', 'time_added': '22:18:53-07:00', 'artist': 'Silverchair', 'artist_id': 18842, 'title': 'Neon Ballroom', 'year': 2010, 'format': 'Vinyl', 'format_info': '180 Gram, Gatefold', 'genres': "'Rock'", 'styles': "'Alternative Rock', 'Hard Rock', 'Grunge'", 'notes': 'Purchased on Amazon, 2017', 'lowest_price': 22.19}, {'id': 6749198, 'master_id': 67090, 'date_added': '2017-06-02', 'time_added': '22:00:10-07:00', 'artist': 'Silverchair', 'artist_id': 18842, 'title': 'Neon Ballroom', 'year': 2015, 'format': 'Vinyl', 'format_info': 'Blue Translucent', 'genres': "'Rock'", 'styles': "'Alternative Rock'", 'notes': "Purchased at Norman's Sound and Vision Records, Brooklyn, 2017", 'lowest_price': 35.0}, {'id': 6841641, 'master_id': 121648, 'date_added': '2017-05-05', 'time_added': '21:38:27-07:00', 'artist': 'UGK', 'artist_id': 133866, 'title': "Ridin' Dirty", 'year': 2015, 'format': 'Vinyl', 'format_info': 'Clear', 'genres': "'Hip Hop'", 'styles': "'Gangsta'", 'notes': 'Purchased at Turntable Lab, Manhattan, RSD 2017', 'lowest_price': 82.42}, {'id': 7492627, 'master_id': 887237, 'date_added': '2017-05-05', 'time_added': '21:57:45-07:00', 'artist': 'Chvrches', 'artist_id': 2953514, 'title': 'Every Open Eye', 'year': 2015, 'format': 'Vinyl', 'format_info': 'Coke-Bottle Translucent, 180 Gram', 'genres': "'Electronic'", 'styles': "'Synth-pop'", 'notes': 'Purchased on Amazon, 2016', 'lowest_price': 19.97}, {'id': 8241863, 'master_id': 0, 'date_added': '2017-05-05', 'time_added': '21:18:41-07:00', 'artist': 'Chvrches', 'artist_id': 2953514, 'title': 'Every Open Eye (The Remixes)', 'year': 2016, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Electronic', 'Pop'", 'styles': "'Synth-pop', 'Techno', 'House'", 'notes': 'Purchased at Bleeker Street Records, Manhattan, RSD 2016', 'lowest_price': 6.5}, {'id': 8967355, 'master_id': 599191, 'date_added': '2017-05-16', 'time_added': '22:15:03-07:00', 'artist': 'Chvrches', 'artist_id': 2953514, 'title': 'The Bones Of What You Believe', 'year': 2016, 'format': 'Vinyl', 'format_info': 'Blue', 'genres': "'Electronic', 'Pop'", 'styles': "'Indie Pop', 'Synth-pop'", 'notes': 'Purchased on Newbury Comics, 2017', 'lowest_price': 74.99}, {'id': 8967415, 'master_id': 887237, 'date_added': '2017-05-22', 'time_added': '18:15:08-07:00', 'artist': 'Chvrches', 'artist_id': 2953514, 'title': 'Every Open Eye', 'year': 2016, 'format': 'Vinyl', 'format_info': 'Pink, 180g', 'genres': "'Electronic', 'Pop'", 'styles': "'Synth-pop'", 'notes': 'Purchased on Newbury Comics, 2017', 'lowest_price': 65.0}, {'id': 8406325, 'master_id': 367797, 'date_added': '2017-05-05', 'time_added': '21:14:08-07:00', 'artist': "D'eon", 'artist_id': 1790730, 'title': 'Darkbloom', 'year': 2016, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Electronic', 'Pop'", 'styles': "'Electro', 'Synth-pop', 'Juke'", 'notes': 'Purchased at Turntable Lab, Manhattan, 2017', 'lowest_price': 14.99}, {'id': 3886754, 'master_id': 470918, 'date_added': '2017-07-10', 'time_added': '16:35:42-07:00', 'artist': 'The Killers', 'artist_id': 220651, 'title': 'Battle Born', 'year': 2012, 'format': 'Vinyl', 'format_info': 'Red, 180 Gram', 'genres': "'Electronic', 'Rock', 'Pop'", 'styles': "'Alternative Rock', 'Pop Rock', 'Synth-pop', 'Indie Rock'", 'notes': 'Purchased on Amazon, 2017', 'lowest_price': 25.0}, {'id': 9642831, 'master_id': 71471, 'date_added': '2017-06-12', 'time_added': '16:23:09-07:00', 'artist': 'The Killers', 'artist_id': 220651, 'title': 'Hot Fuss', 'year': 2017, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Electronic', 'Rock'", 'styles': "'New Wave', 'Synth-pop', 'Indie Rock'", 'notes': 'Purchased on Amazon, 2017', 'lowest_price': 26.68}, {'id': 2590670, 'master_id': 41160, 'date_added': '2017-05-05', 'time_added': '15:23:40-07:00', 'artist': 'Ozzy Osbourne', 'artist_id': 59770, 'title': 'Diary Of A Madman', 'year': 1981, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': "'Heavy Metal'", 'notes': 'Purchased at A1 Record Shop, Manhattan, 2017', 'lowest_price': 12.0}, {'id': 2183941, 'master_id': 41178, 'date_added': '2017-05-05', 'time_added': '18:13:13-07:00', 'artist': 'Ozzy Osbourne', 'artist_id': 59770, 'title': 'Speak Of The Devil', 'year': 1982, 'format': 'Vinyl', 'format_info': 'Carrollton Pressing, Gatefold', 'genres': "'Rock'", 'styles': "'Hard Rock', 'Heavy Metal'", 'notes': 'Purchased at A1 Record Shop, Manhattan, 2017', 'lowest_price': 22.0}, {'id': 7528533, 'master_id': 10236, 'date_added': '2017-05-21', 'time_added': '20:32:07-07:00', 'artist': 'The Sword', 'artist_id': 569308, 'title': 'Age Of Winters', 'year': 2015, 'format': 'Vinyl', 'format_info': 'Purple Marbled', 'genres': "'Rock'", 'styles': "'Stoner Rock', 'Doom Metal'", 'notes': 'Purchased on Newbury Comics, 2017', 'lowest_price': 70.59}, {'id': 8385076, 'master_id': 462621, 'date_added': '2017-05-05', 'time_added': '21:24:57-07:00', 'artist': 'Clutch', 'artist_id': 289116, 'title': 'The Elephant Riders', 'year': 2016, 'format': 'Vinyl', 'format_info': 'Red Translucent, Gatefold', 'genres': "'Rock'", 'styles': "'Alternative Rock', 'Hard Rock', 'Stoner Rock'", 'notes': 'Purchased at Bleeker Street Records, Manhattan, RSD 2016', 'lowest_price': 64.71}, {'id': 1631167, 'master_id': 14563, 'date_added': '2020-10-09', 'time_added': '16:08:25-07:00', 'artist': 'Tangerine Dream', 'artist_id': 10343, 'title': 'Legend (Music From The Motion Picture Soundtrack)', 'year': 1986, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Electronic', 'Stage & Screen'", 'styles': "'Modern Classical', 'Score', 'Synth-pop', 'Ambient'", 'notes': 'Very Good Plus (VG+)', 'lowest_price': 39.5}, {'id': 8070507, 'master_id': 7795, 'date_added': '2017-05-29', 'time_added': '11:57:46-07:00', 'artist': 'Slayer', 'artist_id': 18845, 'title': 'Hell Awaits', 'year': 2016, 'format': 'Vinyl', 'format_info': 'White, 180 Gram', 'genres': "'Rock'", 'styles': "'Thrash'", 'notes': 'Purchased on Nuclear Blast, 2017', 'lowest_price': 0.0}, {'id': 4702208, 'master_id': 35071, 'date_added': '2017-05-05', 'time_added': '21:54:44-07:00', 'artist': 'Weezer', 'artist_id': 105730, 'title': 'Pinkerton', 'year': 2013, 'format': 'Vinyl', 'format_info': 'Gatefold', 'genres': "'Rock'", 'styles': "'Alternative Rock', 'Pop Rock'", 'notes': '# 009365\nPurchase at Turntable Lab, Manhattan, RSD 2017', 'lowest_price': 34.37}, {'id': 4847709, 'master_id': 43553, 'date_added': '2017-05-16', 'time_added': '19:54:20-07:00', 'artist': 'Billy Joel', 'artist_id': 137418, 'title': 'An Innocent Man', 'year': 2013, 'format': 'Vinyl', 'format_info': '180 Gram', 'genres': "'Rock', 'Funk / Soul'", 'styles': "'Rock & Roll', 'Pop Rock', 'Doo Wop', 'Soul'", 'notes': '# 000261\nPurchased on Amazon, 2017', 'lowest_price': 71.43}, {'id': 9225892, 'master_id': 35266, 'date_added': '2017-06-13', 'time_added': '16:47:45-07:00', 'artist': 'Weezer', 'artist_id': 105730, 'title': 'Weezer', 'year': 2016, 'format': 'Vinyl', 'format_info': 'Blue Marbled, 180g, Gatefold', 'genres': "'Rock'", 'styles': "'Alternative Rock', 'Indie Rock'", 'notes': '# 001052\nPurchased on Discogs, 2017', 'lowest_price': 49.99}, {'id': 8399966, 'master_id': 0, 'date_added': '2017-05-05', 'time_added': '21:27:28-07:00', 'artist': 'Anthrax', 'artist_id': 66025, 'title': 'Antisocial (Live) b/w In The End (Live)', 'year': 2016, 'format': 'Vinyl', 'format_info': 'Red', 'genres': "'Rock'", 'styles': "'Thrash', 'Heavy Metal'", 'notes': 'Purchased at Bleeker Street Records, Manhattan, RSD 2016', 'lowest_price': 5.0}, {'id': 9673152, 'master_id': 52213, 'date_added': '2017-06-12', 'time_added': '16:13:30-07:00', 'artist': 'Silverchair', 'artist_id': 18842, 'title': 'Frogstomp', 'year': 1995, 'format': 'Vinyl', 'format_info': 'Green', 'genres': "'Rock'", 'styles': "'Alternative Rock', 'Grunge'", 'notes': 'Purchased on eBay, 2017', 'lowest_price': 105.88}, {'id': 113793, 'master_id': 784584, 'date_added': '2017-05-05', 'time_added': '08:52:03-07:00', 'artist': 'Yazoo', 'artist_id': 2713, 'title': 'Only You ', 'year': 1982, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Electronic', 'Pop'", 'styles': "'Synth-pop'", 'notes': 'Purchased at Good Records NYC, Manhattan, 2017', 'lowest_price': 2.35}, {'id': 8898088, 'master_id': 1042632, 'date_added': '2017-06-04', 'time_added': '19:33:39-07:00', 'artist': 'Savoir Adore', 'artist_id': 1176610, 'title': 'The Love That Remains', 'year': 2016, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Electronic', 'Pop'", 'styles': "'Indie Pop'", 'notes': "Purchased at Norman's Sound and Vision Records, Brooklyn, 2017", 'lowest_price': 8.43}, {'id': 10174334, 'master_id': 17696, 'date_added': '2017-05-05', 'time_added': '21:37:32-07:00', 'artist': 'The Cinematic Orchestra', 'artist_id': 3384, 'title': 'Ma Fleur', 'year': 2017, 'format': 'Vinyl', 'format_info': 'Blue', 'genres': "'Electronic', 'Jazz', 'Rock'", 'styles': "'Soul-Jazz', 'Future Jazz', 'Downtempo'", 'notes': 'Purchased at Turntable Lab, Manhattan, RSD 2017', 'lowest_price': 58.82}, {'id': 10245525, 'master_id': 1153232, 'date_added': '2017-05-29', 'time_added': '12:15:39-07:00', 'artist': 'Pallbearer', 'artist_id': 1983458, 'title': 'Heartless', 'year': 2017, 'format': 'Vinyl', 'format_info': 'Blue Azure', 'genres': "'Rock'", 'styles': "'Doom Metal'", 'notes': 'Purchased on Nuclear Blast, 2017', 'lowest_price': 41.06}, {'id': 7241731, 'master_id': 863514, 'date_added': '2017-05-27', 'time_added': '17:14:31-07:00', 'artist': 'Death Angel', 'artist_id': 266100, 'title': 'The Bay Calls For Blood (Live In San Francisco)', 'year': 2015, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': "'Thrash'", 'notes': 'Purchased on Nuclear Blast, 2017', 'lowest_price': 12.93}, {'id': 8574146, 'master_id': 1005867, 'date_added': '2017-05-31', 'time_added': '10:50:15-07:00', 'artist': 'Death Angel', 'artist_id': 266100, 'title': 'The Evil Divide', 'year': 2016, 'format': 'Vinyl', 'format_info': 'Grey', 'genres': "'Rock'", 'styles': "'Heavy Metal', 'Thrash'", 'notes': 'Purchased on Nuclear Blast, 2017', 'lowest_price': 55.0}, {'id': 8391747, 'master_id': 1929, 'date_added': '2017-05-05', 'time_added': '21:30:37-07:00', 'artist': 'David Bowie', 'artist_id': 10263, 'title': 'The Man Who Sold The World', 'year': 2016, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': "'Hard Rock', 'Glam'", 'notes': 'Purchased at Bleeker Street Records, Manhattan, RSD 2016', 'lowest_price': 18.0}, {'id': 9431705, 'master_id': 10370, 'date_added': '2017-05-19', 'time_added': '19:38:26-07:00', 'artist': 'Pink Floyd', 'artist_id': 45467, 'title': 'Animals', 'year': 2016, 'format': 'Vinyl', 'format_info': 'Gatefold, 180 Gram', 'genres': "'Rock'", 'styles': "'Prog Rock', 'Psychedelic Rock', 'Classic Rock'", 'notes': 'Purchased on Amazon, 2017', 'lowest_price': 19.08}, {'id': 2514427, 'master_id': 58387, 'date_added': '2017-05-05', 'time_added': '21:29:46-07:00', 'artist': 'Roxy Music', 'artist_id': 56621, 'title': 'The Same Old Scene', 'year': 1980, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': "'Pop Rock'", 'notes': 'Free gift at Bleeker Street Records, Manhattan, RSD 2016', 'lowest_price': 0.65}, {'id': 14871374, 'master_id': 28859, 'date_added': '2020-10-06', 'time_added': '14:53:24-07:00', 'artist': 'Cyndi Lauper', 'artist_id': 29718, 'title': "She's So Unusual", 'year': 1983, 'format': 'Vinyl', 'format_info': 'Carrollton Pressing, "Barry" and "Chet" version', 'genres': "'Electronic', 'Rock', 'Pop'", 'styles': "'Pop Rock', 'New Wave'", 'notes': 'Near Mint (NM or M-)', 'lowest_price': 12.5}, {'id': 10234021, 'master_id': 1174493, 'date_added': '2017-05-16', 'time_added': '19:59:23-07:00', 'artist': 'The Sword', 'artist_id': 569308, 'title': 'Greetings From...', 'year': 2017, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': '', 'notes': 'Purchased at Piranha Records, Round Rock, 2017', 'lowest_price': 8.0}, {'id': 3549641, 'master_id': 432684, 'date_added': '2017-05-05', 'time_added': '22:07:05-07:00', 'artist': 'Mastodon', 'artist_id': 252161, 'title': 'Feistodon', 'year': 2012, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock', 'Pop'", 'styles': '', 'notes': 'Purchased at J&R, Manhattan, RSD 2012', 'lowest_price': 15.0}, {'id': 5745724, 'master_id': 18080, 'date_added': '2017-05-16', 'time_added': '21:20:03-07:00', 'artist': 'Depeche Mode', 'artist_id': 2725, 'title': 'Violator', 'year': 2014, 'format': 'Vinyl', 'format_info': '180-Gram, Gatefold, Rainbo Records Pressing', 'genres': "'Electronic'", 'styles': "'Synth-pop'", 'notes': 'Purchased on Amazon, 2017', 'lowest_price': 28.0}, {'id': 8015159, 'master_id': 947809, 'date_added': '2017-06-06', 'time_added': '20:11:30-07:00', 'artist': 'Beastmaker', 'artist_id': 4631594, 'title': 'You Must Sin', 'year': 2016, 'format': 'Vinyl', 'format_info': 'Purple', 'genres': "'Rock'", 'styles': "'Doom Metal'", 'notes': 'Purchased on Amazon, 2017', 'lowest_price': 8.7}, {'id': 9132029, 'master_id': 987434, 'date_added': '2017-07-11', 'time_added': '17:30:31-07:00', 'artist': 'Beastmaker', 'artist_id': 4631594, 'title': 'Lusus Naturae', 'year': 2016, 'format': 'Vinyl', 'format_info': ' White', 'genres': "'Rock'", 'styles': "'Doom Metal'", 'notes': 'Purchased on Amazon, 2017', 'lowest_price': 41.18}, {'id': 10316860, 'master_id': 1182583, 'date_added': '2017-06-06', 'time_added': '20:13:08-07:00', 'artist': 'Beastmaker', 'artist_id': 4631594, 'title': 'Inside The Skull', 'year': 2017, 'format': 'Vinyl', 'format_info': 'Purple', 'genres': "'Rock'", 'styles': "'Doom Metal'", 'notes': 'Purchased on Amazon, 2017', 'lowest_price': 19.47}, {'id': 8397183, 'master_id': 0, 'date_added': '2017-05-05', 'time_added': '21:19:45-07:00', 'artist': 'John Coltrane', 'artist_id': 97545, 'title': 'The Roulette Sides', 'year': 2016, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Jazz'", 'styles': '', 'notes': 'Purchased at Bleeker Street Records, Manhattan, RSD 2016', 'lowest_price': 10.0}, {'id': 8345669, 'master_id': 252597, 'date_added': '2017-06-12', 'time_added': '18:01:56-07:00', 'artist': 'Buddy Holly', 'artist_id': 272426, 'title': 'Buddy', 'year': 2016, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': "'Rock & Roll'", 'notes': 'Purchased on Amazon, 2017', 'lowest_price': 11.76}, {'id': 7287188, 'master_id': 67087, 'date_added': '2017-06-13', 'time_added': '16:40:26-07:00', 'artist': 'Silverchair', 'artist_id': 18842, 'title': 'Freak Show', 'year': 2015, 'format': 'Vinyl', 'format_info': 'Red, 180 Gram', 'genres': "'Rock'", 'styles': '', 'notes': 'Purchased on eBay, 2017', 'lowest_price': 50.0}, {'id': 10203377, 'master_id': 1138833, 'date_added': '2017-06-11', 'time_added': '14:03:35-07:00', 'artist': 'Power Trip', 'artist_id': 2408820, 'title': 'Nightmare Logic', 'year': 2017, 'format': 'Vinyl', 'format_info': 'Purple', 'genres': "'Rock'", 'styles': "'Hardcore', 'Thrash'", 'notes': 'Purchased on Southern Lord, 2017', 'lowest_price': 0.0}, {'id': 10318718, 'master_id': 1185560, 'date_added': '2017-05-19', 'time_added': '20:28:32-07:00', 'artist': 'Zakk Sabbath', 'artist_id': 5770577, 'title': 'Live In Detroit', 'year': 2017, 'format': 'Vinyl', 'format_info': 'Purple', 'genres': "'Rock'", 'styles': "'Heavy Metal'", 'notes': 'Purchased at live concert, Gramercy Theatre, Manhattan, 5/18/17', 'lowest_price': 0.0}, {'id': 10357365, 'master_id': 1185560, 'date_added': '2017-06-11', 'time_added': '11:12:58-07:00', 'artist': 'Zakk Sabbath', 'artist_id': 5770577, 'title': 'Live In Detroit', 'year': 2017, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': "'Heavy Metal'", 'notes': 'Purchased on Southern Lord, 2017', 'lowest_price': 0.0}, {'id': 8212197, 'master_id': 564069, 'date_added': '2017-06-11', 'time_added': '11:16:24-07:00', 'artist': 'Power Trip', 'artist_id': 2408820, 'title': 'Manifest Decimation', 'year': 2013, 'format': 'Vinyl', 'format_info': 'Clear', 'genres': "'Rock'", 'styles': "'Thrash', 'Hardcore'", 'notes': 'Purchased on Southern Lord, 2017', 'lowest_price': 0.0}, {'id': 6737156, 'master_id': 67090, 'date_added': '2017-06-12', 'time_added': '15:32:57-07:00', 'artist': 'Silverchair', 'artist_id': 18842, 'title': 'Neon Ballroom', 'year': 2015, 'format': 'Vinyl', 'format_info': 'Pink', 'genres': "'Rock'", 'styles': "'Alternative Rock'", 'notes': 'Purchased on eBay, 2017', 'lowest_price': 0.0}, {'id': 6761444, 'master_id': 67090, 'date_added': '2017-07-10', 'time_added': '17:00:36-07:00', 'artist': 'Silverchair', 'artist_id': 18842, 'title': 'Neon Ballroom', 'year': 2015, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': "'Alternative Rock'", 'notes': 'Purchased on Discogs, 2017', 'lowest_price': 0.0}, {'id': 8014535, 'master_id': 946104, 'date_added': '2017-05-05', 'time_added': '20:00:05-07:00', 'artist': 'Megadeth', 'artist_id': 11770, 'title': 'Dystopia', 'year': 2016, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': "'Heavy Metal', 'Thrash'", 'notes': 'Free LP with Target CD Preorder, 2016', 'lowest_price': 16.0}, {'id': 8627508, 'master_id': 17217, 'date_added': '2017-06-12', 'time_added': '17:15:36-07:00', 'artist': 'The Beach Boys', 'artist_id': 70829, 'title': 'Pet Sounds', 'year': 2016, 'format': 'Vinyl', 'format_info': '200 Gram', 'genres': "'Rock', 'Pop'", 'styles': "'Psychedelic Rock', 'Pop Rock'", 'notes': 'Purchased at Crate & Barrel, 2017', 'lowest_price': 18.82}, {'id': 1903476, 'master_id': 11019, 'date_added': '2017-07-11', 'time_added': '15:41:08-07:00', 'artist': 'Muse', 'artist_id': 1003, 'title': 'Origin Of Symmetry', 'year': 2009, 'format': 'Vinyl', 'format_info': 'Gatefold', 'genres': "'Rock'", 'styles': "'Alternative Rock', 'Prog Rock'", 'notes': 'Purchased on Insound, 2017', 'lowest_price': 15.0}, {'id': 1893881, 'master_id': 11052, 'date_added': '2017-06-08', 'time_added': '21:11:22-07:00', 'artist': 'Muse', 'artist_id': 1003, 'title': 'Absolution', 'year': 2009, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': "'Alternative Rock'", 'notes': 'Purchased on InSound, 2017', 'lowest_price': 17.65}, {'id': 1660254, 'master_id': 85129, 'date_added': '2017-06-08', 'time_added': '19:04:12-07:00', 'artist': 'Eric Clapton', 'artist_id': 17827, 'title': 'August', 'year': 1986, 'format': 'Vinyl', 'format_info': 'Gatefold', 'genres': "'Rock'", 'styles': '', 'notes': 'Purchased at A1 Record Shop, Manhattan, 2017', 'lowest_price': 2.06}, {'id': 7088529, 'master_id': 842235, 'date_added': '2017-07-11', 'time_added': '17:33:58-07:00', 'artist': 'Muse', 'artist_id': 1003, 'title': 'Drones', 'year': 2015, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': "'Alternative Rock'", 'notes': 'Purchased on Insound, 2017', 'lowest_price': 9.41}, {'id': 8228904, 'master_id': 1377759, 'date_added': '2017-05-05', 'time_added': '21:26:29-07:00', 'artist': 'Muse', 'artist_id': 1003, 'title': 'Reapers', 'year': 2016, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': "'Alternative Rock', 'Hard Rock'", 'notes': 'Purchased at Bleeker Street Records, Manhattan, RSD 2016', 'lowest_price': 4.55}, {'id': 3919402, 'master_id': 475177, 'date_added': '2017-06-11', 'time_added': '10:40:10-07:00', 'artist': 'Muse', 'artist_id': 1003, 'title': 'The 2nd Law', 'year': 2012, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': "'Alternative Rock'", 'notes': 'Purchased on InSound, 2017', 'lowest_price': 24.99}, {'id': 8867282, 'master_id': 4785, 'date_added': '2017-07-11', 'time_added': '16:35:24-07:00', 'artist': 'Black Sabbath', 'artist_id': 144998, 'title': 'Sabotage', 'year': 2016, 'format': 'Vinyl', 'format_info': 'Purple, 180g', 'genres': "'Rock'", 'styles': "'Hard Rock', 'Heavy Metal'", 'notes': 'Purchased on Discogs, 2017', 'lowest_price': 69.41}, {'id': 3554343, 'master_id': 404501, 'date_added': '2017-05-05', 'time_added': '22:04:45-07:00', 'artist': 'Metallica', 'artist_id': 18839, 'title': 'Beyond Magnetic', 'year': 2012, 'format': 'Vinyl', 'format_info': 'Silver', 'genres': "'Rock'", 'styles': "'Thrash', 'Heavy Metal'", 'notes': 'Purchased at J&R, Manhattan, RSD 2012', 'lowest_price': 62.62}, {'id': 4388548, 'master_id': 29419, 'date_added': '2017-05-05', 'time_added': '20:57:35-07:00', 'artist': 'Van Halen', 'artist_id': 94066, 'title': '1984', 'year': 1984, 'format': 'Vinyl', 'format_info': 'Columbia House', 'genres': "'Rock'", 'styles': "'Hard Rock'", 'notes': 'Purchased at A1 Record Shop, Manhattan, 2017', 'lowest_price': 21.99}, {'id': 7883349, 'master_id': 930482, 'date_added': '2017-05-24', 'time_added': '05:34:24-07:00', 'artist': 'Clutch', 'artist_id': 289116, 'title': 'Psychic Warfare', 'year': 2015, 'format': 'Vinyl', 'format_info': 'Red Translucent', 'genres': "'Rock'", 'styles': "'Hard Rock'", 'notes': 'Purchased on Newbury Comics, 2017', 'lowest_price': 50.0}, {'id': 8388184, 'master_id': 0, 'date_added': '2017-05-05', 'time_added': '21:25:42-07:00', 'artist': 'Clutch', 'artist_id': 289116, 'title': 'Mad Sidewinder / Outland Special Clearance', 'year': 2016, 'format': 'Vinyl', 'format_info': 'N/A', 'genres': "'Rock'", 'styles': "'Stoner Rock', 'Hard Rock'", 'notes': '# 4235/5000\nPurchased at Bleeker Street Records, Manhattan, RSD 2016', 'lowest_price': 9.99}]
